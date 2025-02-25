from asyncio.windows_events import NULL
import string
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from SgpProcessor import SgpProcessor


class SgpHitters():
    def __init__(self,proj,player_sheet:string) -> None:
        wb = load_workbook('LeagueStatsSGPInvest.xlsm', data_only=True)
        sheet = wb["3YR RUNNING AVG SGP"]
        sheet2 = wb[player_sheet]

        self.team_opportunities = {
            'AB' : sheet2['AV2'].value, 
            'PA' : sheet2['AP2'].value
        }

        self.team_value = {
            'OBP_PA' : sheet2['AQ2'].value,
            'SLUG_AB' : sheet2['AW2'].value
        }

        self.replacement_levels = {
            'R': sheet['Q26'].value, 'HR': sheet['R26'].value, 'RBI': sheet['S26'].value, 'SB': sheet['T26'].value,
            'OBP': sheet['U26'].value, 'SLG': sheet['V26'].value
        }
    
        self.cat_stds = {
            'R': sheet['Q27'].value, 'HR': sheet['R27'].value, 'RBI': sheet['S27'].value, 'SB': sheet['T27'].value,
            'OBP': sheet['U27'].value, 'SLG': sheet['V27'].value
        }
        
        self.stats = pd.read_excel('fangraphs_hitting_atc.xlsx', sheet_name=0)
        self.process_hitters_sgp()
        self.sgp_df[['Name', 'PlayerId']] = self.stats[['Name', 'PlayerId']]
        self.sgp_df.set_index(['Name','PlayerId'], inplace=True)

    def cat_calc_sgp(self,projection:string):
        return (self.stats[projection] - self.replacement_levels[projection]) / self.cat_stds[projection]

    def rate_calc_sgp(self,cat:string,opportunities:string):
        if opportunities == 'AB':
            team_cat = 'SLUG_AB'
            player_opps = self.stats[opportunities]
        elif opportunities == 'PA':
            team_cat = 'OBP_PA'
            player_opps = self.stats[opportunities] - self.stats['SH']

        player_val = self.stats[cat]*player_opps
        team_val_wo_average_player = self.team_value[team_cat]
        total_opps = self.team_opportunities[opportunities] + player_opps
        return ((team_val_wo_average_player+player_val)/(total_opps) - self.replacement_levels[cat])/self.cat_stds[cat]

    def process_hitters_sgp(self):
        self.sgp_df = pd.DataFrame()
        for cat in ['R', 'HR', 'RBI', 'SB']:
            self.sgp_df[f'SGP_{cat}'] = self.cat_calc_sgp(cat)
        
        for cat, opps in [('OBP', 'PA'), ('SLG', 'AB')]:
            self.sgp_df[f'SGP_{cat}'] = self.rate_calc_sgp(cat,opps)
        
            
        

class SgpPitchers():
    def __init__(self,proj:string,player_sheet:string):
        wb = load_workbook('LeagueStatsSGPInvest.xlsm', data_only=True)
        sheet = wb["3YR RUNNING AVG SGP"]
        sheet2 = wb[player_sheet]
        
        self.team_opportunities = {
            'IP' : sheet2['AR2'].value,
            'BB': sheet2['BA2'].value
        }
        self.team_value = {
            'ERA': sheet2['AS2'].value,
            'WHIP': sheet2['AV2'].value,
            'K/BB' : sheet2['BB2'].value

        }

        self.replacement_levels = {
            'SO': sheet['W26'].value,'QS': sheet['X26'].value, 'SV_HLD': sheet['AB26'].value, 
            'ERA': sheet['Y26'].value,'WHIP': sheet['Z26'].value, 'K/BB': sheet['AA26'].value
        }
    
        self.cat_stds = {
            'SO': sheet['W27'].value,'QS': sheet['X27'].value, 'SV_HLD': sheet['AB27'].value, 
            'ERA': sheet['Y27'].value, 'WHIP': sheet['Z27'].value, 'K/BB': sheet['AA27'].value
        }
        
        self.stats = pd.read_excel('fangraphs_pitching_atc.xlsx', sheet_name=0)
        self.process_pitchers_sgp()
        self.sgp_df[['Name', 'PlayerId']] = self.stats[['Name', 'PlayerId']]
        self.sgp_df.set_index(['Name','PlayerId'], inplace=True)

    def cat_calc_sgp(self,projection,cat:string):
        return (projection - self.replacement_levels[cat]) / self.cat_stds[cat]

    def pitcher_rate_calc(self,projection,cat,opps):
        multiplier = 1
        if (cat == 'ERA'):
            multiplier = 9
        team_val_wo_average_player = multiplier*self.team_value[cat]
        total_opps = self.team_opportunities[opps] + self.stats[opps]
        return ((team_val_wo_average_player+projection)/(total_opps) - self.replacement_levels[cat])/self.cat_stds[cat]

    def rate_calc_sgp(self,cat,opps):
        if(cat=='ERA'):
            val = 9*self.stats['ER']
        elif(cat=='WHIP'):
            val = self.stats['H']+self.stats['BB']
        elif(cat=="K/BB"):
            val = self.stats['SO']
        else:
            return NULL
        return self.pitcher_rate_calc(val,cat,opps)

    def process_pitchers_sgp(self):
        self.sgp_df = pd.DataFrame()
        for cat in ['SO', 'QS', 'SV_HLD']:
            if cat == 'SV_HLD':
                val = self.stats['SV'] + self.stats['HLD']
            else:
                val = self.stats[cat]
            self.sgp_df[f'SGP_{cat}'] = self.cat_calc_sgp(val,cat)
        
        for cat, opps in [('ERA','IP'), ('WHIP', 'IP'), ('K/BB', 'BB')]:
            self.sgp_df[f'SGP_{cat}'] = self.rate_calc_sgp(cat,opps)
        
    
if __name__ == "__main__":
    sgp_hit = SgpHitters(proj="ATC HIT '25",player_sheet="SGP ATC HIT '25")
    sgp_pit = SgpPitchers(proj="ATC PIT '25",player_sheet="SGP ATC PIT '25")
    #sgp_pit_oopsy = SgpPitchers(proj="SGP PIT OOPSY '25",player_sheet="SGP OOPSY PIT '25")
    
    processor = SgpProcessor(sgp_hit,sgp_pit)    
    

